from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfile

#encodes number for excel columns
def calc_col(n):
    div = int(n / 26)
    mod = n % 26
    if(div > 0):
        return calc_col(div-1)+chr(mod+65)
    if(div == 0):
        return chr(mod+65)

#keep proportion using the large side resolution to the large side
def keep_proportion(large_side, width, height):
    if (width > height):
        return (large_side, int((large_side / width) * height))
    else:
        return (int((large_side / height) * width), large_side)

#const with large side resolution
LARGE_SIDE_RESOLUTION = 200

Tk().withdraw()                 # we don't want a full GUI, so keep the root window from appearing
                                # show an "Open" dialog box and return the path to the selected file
filename = askopenfilename(
    filetypes=[('PNG', '.png'), ('JPG', '.jpg'), ('JPG', '.jpeg'), ('GIF', '.gif')])

save_name = asksaveasfile(mode='w', filetypes = [('Excel', '.xlsx')], defaultextension='.xlsx')

img = Image.open(filename)
width, height = img.size

#set new width and new height keeping the proportion
n_width, n_height = keep_proportion(LARGE_SIDE_RESOLUTION, width, height)
img = img.resize((n_width, n_height), Image.ANTIALIAS)

#convert image to RGB
rgb_im = img.convert('RGB')
excel_file = Workbook()
sheet = excel_file.active
#scrolls pixel by pixel to fill the table
for i in range(1, n_width):
    sheet.column_dimensions[calc_col(i-1)].width = 3
    for j in range(1, n_height):
        #convert RGB to HEX format color
        string = '%02x%02x%02x' % rgb_im.getpixel((i, j))
        #set the cell with the correspondent pixel color
        sheet.cell(row=j, column=i).fill = PatternFill(
            start_color=string, end_color=string, fill_type="solid")
#save exel file
excel_file.save(save_name.name)
